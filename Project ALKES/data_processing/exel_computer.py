# data_processing/excel_computer.py

import openpyxl
import re
from .logger import Logger

class Computer:

    # Варианты статуса
    STATUS_OPTIONS = [
        "Ок",
        "Не согласовано/Защита отключена",
        "Не согласовано/Обнаружен браузер",
        "Не согласовано/Давно не был в сети",
        ""
    ]

    # Варианты браузера
    BROWSER_OPTIONS = [
        "Chrome",
        "Firefox",
        "Opera",
        ""
    ]

    def __init__(self, position: int, pc_name: str):
        self.position = position  # Позиция в таблице
        self.pc_name = pc_name     # Имя ПК
        self.status = ""        # Статус по умолчанию
        self.browser = ""  # Браузер по умолчанию

    def set_status(self, status: str):
        """Метод для установки статуса, если он в допустимых вариантах."""
        if status in self.STATUS_OPTIONS:
            self.status = status
            Logger.log_info(f"Статус ПК '{self.pc_name}' изменён на '{status}'.")
        else:
            Logger.log_error(f"Неверный статус '{status}' для ПК '{self.pc_name}'. Доступные варианты: {self.STATUS_OPTIONS}")
            raise ValueError(f"Неверный статус. Доступные варианты: {self.STATUS_OPTIONS}")

    def set_browser(self, browser: str):
        """Метод для установки браузера, если он в допустимых вариантах."""
        if browser in self.BROWSER_OPTIONS:
            self.browser = browser
            Logger.log_info(f"Браузер ПК '{self.pc_name}' изменён на '{browser}'.")
        else:
            Logger.log_error(f"Неверный браузер '{browser}' для ПК '{self.pc_name}'. Доступные варианты: {self.BROWSER_OPTIONS}")
            raise ValueError(f"Неверный браузер. Доступные варианты: {self.BROWSER_OPTIONS}")

    def __str__(self):
        return f"ПК: {self.pc_name}, Позиция: {self.position}, Статус: {self.status}, Браузер: {self.browser}"

    @classmethod
    def load_computers(cls, filename, create_files=False):
        """Метод для загрузки объектов Computer из Excel файла и создания файлов MIR и MSK, если указано."""
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active  # Или укажите нужный лист по имени
        
        # Извлекаем заголовки из первой строки
        headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

        # Находим индексы нужных колонок
        try:
            pc_name_col_index = headers.index("ИМЯ ПК") + 1  # +1 для индексации в Excel
            site_col_index = headers.index("Площадка") + 1  # индекс для Площадки
        except ValueError as e:
            # Логируем ошибку через Logger
            Logger.log_error(f"Ошибка: {str(e)}")
            return {}  # Если столбец не найден, возвращаем пустой словарь

        computers = {}  # Словарь для хранения объектов Computer
        mir_computers = []  # Список для имен компьютеров для MIR
        msk_computers = []  # Список для имен компьютеров для MSK

        # Проходим по всем строкам, начиная со второй
        for row in range(2, sheet.max_row + 1):
            pc_name = sheet.cell(row, pc_name_col_index).value
            site_value = sheet.cell(row, site_col_index).value  # Значение в столбце "Площадка"

            if pc_name:  # Проверяем, что имя ПК не пустое
                # Создаем экземпляр Computer и добавляем в словарь
                computer = cls(row, pc_name)  # Используем номер строки как позицию
                computers[pc_name] = computer  # Добавляем в словарь по имени ПК

                # Добавляем имя компьютера в соответствующий список
                if site_value == "MIR":
                    mir_computers.append(pc_name)
                elif site_value == "MSK":
                    msk_computers.append(pc_name)

        # Извлечение даты из имени файла
        match = re.search(r"\d{4}[-_]\d{2}[-_]\d{2}", filename)
        extract_date_value = match.group(0) if match else None
        
        if not extract_date_value:
            Logger.log_error(f"Дата не найдена в названии файла: {filename}. Файлы не будут созданы.")
            return computers  # Возвращаем загруженные компьютеры, даже если файлы не созданы

        if create_files:
            mir_filename = f"{extract_date_value}_MIR.txt"
            msk_filename = f"{extract_date_value}_MSK.txt"

            with open(mir_filename, "w", encoding="utf-8") as mir_file:
                for computer in mir_computers:
                    mir_file.write(computer + "\n")

            with open(msk_filename, "w", encoding="utf-8") as msk_file:
                for computer in msk_computers:
                    msk_file.write(computer + "\n")

            # Запись в лог-файл о создании файлов
            Logger.log_info(f"Созданы файлы: {mir_filename} и {msk_filename} для компьютеров из {filename}.")
        
        # Записываем количество загруженных компьютеров в лог
        Logger.log_info(f"Загружено {len(computers)} компьютеров.")
        return computers  # Возвращаем словарь загруженных компьютеров