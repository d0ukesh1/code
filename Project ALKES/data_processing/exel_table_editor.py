# data_processing/excel_table_editor.py

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from .logger import Logger

class TableEditor:
    def __init__(self, file_name: str, sheet_name: str = None):
        """
        Инициализирует TableEditor с указанным файлом Excel и листом.
        :param file_name: Имя файла Excel.
        :param sheet_name: Имя листа в книге (по умолчанию активный лист).
        """
        self.file_name = file_name
        try:
            # Загружаем книгу
            self.workbook = openpyxl.load_workbook(file_name, data_only=True)
            # Выбираем указанный лист или активный лист по умолчанию
            # Если sheet_name не указан, берем первый лист
            if sheet_name is None:
                self.sheet = self.workbook.active  # Получаем активный (первый) лист
            else:
                self.sheet = self.workbook[sheet_name]  # Получаем указанный лист
        except FileNotFoundError:
            Logger.log_error(f"Файл не найден: {file_name}")
        except Exception as e:
            Logger.log_error(f"Ошибка при открытии файла: {str(e)}")


    # Метод для удаления пустых строк и удаления пробелов в именах ПК
    # Метод для удаления строк по заданным условиям и удаления пробелов в именах ПК
    def default_edit(self):
        """Метод для удаления строк по условиям и удаления пробелов в именах ПК."""
        # Извлекаем заголовки из первой строки
        headers = [self.sheet.cell(row=1, column=col).value for col in range(1, self.sheet.max_column + 1)]
        
        # Находим индекс столбца "Имя ПК"
        try:
            pc_name_col_index = headers.index("ИМЯ ПК") + 1  # +1 для соответствия с индексом в Excel
            Logger.log_info("Столбец 'Имя ПК' найден.")
        except ValueError:
            Logger.log_error("Столбец 'Имя ПК' не найден.")
            return  # Если столбец не найден, выходим из метода

        # Удаляем строки, если в 4, 5 или 6 столбцах есть данные, а в 1 и 9 - нет
        for row in range(self.sheet.max_row, 1, -1):  # Идем с конца к началу
            col_1_value = self.sheet.cell(row, 1).value
            col_4_value = self.sheet.cell(row, 4).value
            col_5_value = self.sheet.cell(row, 5).value
            col_6_value = self.sheet.cell(row, 6).value
            col_9_value = self.sheet.cell(row, 9).value
            
            # Проверяем условия для удаления строки
            if (col_4_value or col_5_value or col_6_value) and not (col_1_value or col_9_value):
                self.sheet.delete_rows(row)
                Logger.log_info(f"Удалена строка {row} (данные в 4, 5 или 6 столбцах, но нет в 1 и 9).")

        # Удаляем пробелы в названиях ПК в столбце "Имя ПК"
        for row in range(2, self.sheet.max_row + 1):
            pc_name = self.sheet.cell(row, pc_name_col_index).value  # Используем индекс столбца Имя ПК
            if pc_name and isinstance(pc_name, str):
                new_pc_name = pc_name.replace(" ", "")  # Удаляем пробелы
                self.sheet.cell(row, pc_name_col_index).value = new_pc_name  # Записываем новое имя ПК в ячейку
                if pc_name != new_pc_name:  # Если имя изменилось
                    Logger.log_info(f"Пробелы удалены в строке {row}: '{pc_name}' -> '{new_pc_name}'")  # Логируем изменение

        Logger.log_info("Удаление пробелов завершено.")
        self.format_table()
        Logger.log_info("Форматирование завершено.")
        # Сохраняем изменения в файле
        try:
            self.workbook.save(self.file_name)
            Logger.log_info(f"Изменения сохранены в {self.file_name}.")
        except Exception as e:
            Logger.log_error(f"Ошибка при сохранении файла {self.file_name}: {e}")


    def add_column_if_missing(self, column_name: str):
        """Проверяет имена столбцов и добавляет нужный в конец столбцов, если отсутствует искомый."""
        # Получаем все значения из первой строки (заголовок таблицы)
        if column_name not in [cell.value for cell in self.sheet[1]]:
            new_column_index = self.sheet.max_column + 1

            # Копируем стиль шрифта из первой ячейки заголовка
            existing_font = self.sheet.cell(row=1, column=1).font

            # Создаем шрифт для нового заголовка
            new_font = Font(
                name=existing_font.name,
                size=existing_font.size,
                bold=existing_font.bold or False,
                italic=existing_font.italic,
                underline=existing_font.underline,
                color=existing_font.color
            )
            
            self.sheet.cell(row=1, column=new_column_index, value=column_name).font = new_font
            Logger.log_info(f"Столбец {column_name} добавлен как столбец номер {new_column_index}")

    def fill_cell(self, row: int, column: int, value):
        """Заполняет данные в нужной ячейке и столбце."""
        self.sheet.cell(row=row, column=column, value=value)
        self.workbook.save(self.file_name)  # Сохраняем изменения

    def find_elements_in_checkfile(self, data: list):
        """
        Ищет элементы в первом столбце первого листа файла и добавляет их в массив, если они найдены в списке data.
        """
        found_elements = []  # Список для хранения найденных элементов
        self.sheet = self.workbook.worksheets[0]
        # Получаем первый лист
        if self.sheet:
            # Проходим по всем строкам в первом столбце, начиная со второй (первая строка - заголовки)
            for row in range(1, self.sheet.max_row + 1):
                cell_value = self.sheet.cell(row=row, column=1).value  # Значение из первого столбца
                if cell_value in data:  # Проверяем, есть ли значение в списке data
                    found_elements.append(cell_value)  # Добавляем найденное значение в список

        return found_elements

    def format_table(self):
        """Форматирует всю таблицу."""

        # Задаем размеры ячеек
        heights = [int(190 * 0.75)] + [None] * 12  # Высота заголовков в пунктах
        widths = [int(300 / 8), int(200 / 8), int(200 / 8), int(110 / 8),
                int(210 / 8), int(300 / 8), int(300 / 8), int(125 / 8),
                int(150 / 8), int(290 / 8), int(110 / 8), int(150 / 8),
                int(150 / 8)]  # Ширина столбцов в единицах openpyxl

        # Устанавливаем высоту для заголовков
        for index, height in enumerate(heights, start=1):
            if height is not None:
                self.sheet.row_dimensions[index].height = height

        # Устанавливаем ширину столбцов
        for index, width in enumerate(widths, start=1):
            self.sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

        # Устанавливаем границы для всех ячеек
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for row in self.sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Изменяем высоту заполненных строк до 23 пикселей
        for row in self.sheet.iter_rows(min_row=2):  # Пропускаем заголовки
            if any(cell.value for cell in row):  # Проверяем, заполнена ли хотя бы одна ячейка в строке
                self.sheet.row_dimensions[row[0].row].height = 18  # Устанавливаем высоту строки
        
        # Устанавливаем выравнивание по центру и перенос текста для всех ячеек
        for row in self.sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        

    def highlight_text_in_red(self, row: int):
        """Окрашивает текст в красный цвет в указанной строке, не изменяя других параметров шрифта."""
        for cell in self.sheet[row]:
            # Получаем текущий шрифт для ячейки
            current_font = cell.font
            # Устанавливаем новый шрифт только с измененным цветом
            new_font = Font(color="FF0000", bold=current_font.bold, italic=current_font.italic,
                            name=current_font.name, size=current_font.size, underline=current_font.underline)
            cell.font = new_font  # Применяем новый шрифт

        self.workbook.save(self.file_name)  # Сохраняем изменения
    
    def highlight_text_in_black(self, row: int):
        """Окрашивает текст в черный цвет в указанной строке, не изменяя других параметров шрифта."""
        for cell in self.sheet[row]:
            # Получаем текущий шрифт для ячейки
            current_font = cell.font
            # Устанавливаем новый шрифт только с измененным цветом
            new_font = Font(color="FFFFFF", bold=current_font.bold, italic=current_font.italic,
                            name=current_font.name, size=current_font.size, underline=current_font.underline)
            cell.font = new_font  # Применяем новый шрифт

        self.workbook.save(self.file_name)  # Сохраняем изменения
